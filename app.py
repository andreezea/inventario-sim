"""
Inventario de Chips SIM - App interna de control de inventario
================================================================
Captura fotos con la camara del celular/laptop, detecta el ICCID y el
IMEI/REIF por codigo de barras/QR (pyzbar) o por OCR (pytesseract),
evita duplicados de ICCID, muestra la lista en tiempo real, guarda la
foto de cada chip y permite exportar todo (datos + miniaturas de fotos)
a un Excel.

Incluye login de administrador (unico que puede descargar el Excel), un
selector de tienda que cada usuario elige una vez por sesion, y un panel
de diagnostico para depurar por que no se detecta un codigo.

Ejecutar con:
    streamlit run app.py
"""

import io
import os
import re
import sqlite3
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image, ImageEnhance, ImageOps

# ----------------------------------------------------------------------
# Dependencias opcionales: la app debe seguir funcionando (modo manual)
# aunque falten librerias de sistema (zbar / tesseract) en el equipo.
# ----------------------------------------------------------------------
try:
    from pyzbar.pyzbar import decode as zbar_decode
    ZBAR_AVAILABLE = True
except Exception:
    ZBAR_AVAILABLE = False

try:
    import cv2
    CV2_AVAILABLE = True
except Exception:
    CV2_AVAILABLE = False

try:
    import pytesseract
    TESSERACT_AVAILABLE = True
except Exception:
    TESSERACT_AVAILABLE = False

DB_PATH = "inventario_sim.db"
FOTOS_DIR = "fotos_chips"
# ICCID "clasico": empieza en 89 (industria de telecom), 18-20 digitos en total
ICCID_REGEX = re.compile(r"89\d{16,18}")
# Respaldo: cualquier corrida de 18-20 digitos seguidos (por si el OCR no
# lee bien los primeros dos digitos, o el chip no sigue el prefijo 89).
DIGIT_RUN_REGEX = re.compile(r"\d{18,20}")
THUMB_PX = 90  # tamano de la miniatura embebida en el Excel

# Lista de tiendas para el selector. Edita esta lista con los nombres reales.
TIENDAS = ["Tienda 1", "Tienda 2", "Tienda 3", "Tienda 4", "Tienda 5"]

# Contrasena de administrador: se lee de .streamlit/secrets.toml (clave ADMIN_PASSWORD).
# Nunca queda escrita en este archivo. Ver README para configurarla.
ADMIN_PASSWORD = st.secrets.get("ADMIN_PASSWORD", "")


# ----------------------------------------------------------------------
# Base de datos (SQLite) - persiste entre sesiones y reinicios de la app
# ----------------------------------------------------------------------
@st.cache_resource
def get_connection():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS chips (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            iccid TEXT UNIQUE NOT NULL,
            fecha_hora TEXT NOT NULL,
            metodo TEXT,
            foto_path TEXT,
            tienda TEXT,
            imei TEXT
        )
        """
    )
    # Migracion simple para bases de datos creadas con una version anterior
    # de la app (sin columna foto_path / tienda / imei).
    cols = [row[1] for row in conn.execute("PRAGMA table_info(chips)").fetchall()]
    if "foto_path" not in cols:
        conn.execute("ALTER TABLE chips ADD COLUMN foto_path TEXT")
    if "tienda" not in cols:
        conn.execute("ALTER TABLE chips ADD COLUMN tienda TEXT")
    if "imei" not in cols:
        conn.execute("ALTER TABLE chips ADD COLUMN imei TEXT")
    conn.commit()
    return conn


def get_all(conn):
    return pd.read_sql_query(
        "SELECT iccid AS ICCID, imei AS IMEI, tienda AS Tienda, "
        "fecha_hora AS 'Fecha y Hora de Captura', metodo AS Metodo, "
        "foto_path AS FotoPath "
        "FROM chips ORDER BY id DESC",
        conn,
    )


def iccid_exists(conn, iccid):
    cur = conn.execute("SELECT 1 FROM chips WHERE iccid = ?", (iccid,))
    return cur.fetchone() is not None


def insert_chip(conn, iccid, metodo, foto_path, tienda, imei):
    conn.execute(
        "INSERT INTO chips (iccid, fecha_hora, metodo, foto_path, tienda, imei) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (iccid, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), metodo, foto_path, tienda, imei),
    )
    conn.commit()


def delete_chip(conn, iccid):
    row = conn.execute("SELECT foto_path FROM chips WHERE iccid = ?", (iccid,)).fetchone()
    conn.execute("DELETE FROM chips WHERE iccid = ?", (iccid,))
    conn.commit()
    if row and row[0] and os.path.exists(row[0]):
        try:
            os.remove(row[0])
        except OSError:
            pass


def save_photo(iccid, img_bytes):
    """Guarda la foto en disco (comprimida) asociada al ICCID. Devuelve la ruta o None si falla."""
    try:
        os.makedirs(FOTOS_DIR, exist_ok=True)
        image = Image.open(io.BytesIO(img_bytes)).convert("RGB")
        image.thumbnail((1000, 1000))
        path = os.path.join(FOTOS_DIR, f"{iccid}.jpg")
        image.save(path, "JPEG", quality=75)
        return path
    except Exception:
        return None


# ----------------------------------------------------------------------
# Deteccion de codigos: ICCID e IMEI/REIF, por codigo de barras/QR y OCR
# ----------------------------------------------------------------------
def _normalize_image(image_pil):
    """Corrige la orientacion (segun metadatos EXIF de la camara) y mejora
    contraste/nitidez antes de intentar leer el codigo de barras o el OCR.
    No modifica la foto que se guarda como respaldo, solo la copia usada
    para la deteccion."""
    try:
        image_pil = ImageOps.exif_transpose(image_pil)
    except Exception:
        pass
    try:
        image_pil = ImageEnhance.Contrast(image_pil).enhance(1.3)
        image_pil = ImageEnhance.Sharpness(image_pil).enhance(1.5)
    except Exception:
        pass
    return image_pil


def _preprocess_for_ocr(image_pil):
    img = np.array(image_pil.convert("L"))
    if CV2_AVAILABLE:
        img = cv2.GaussianBlur(img, (3, 3), 0)
        _, img = cv2.threshold(img, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    return img


def collect_candidates(image_pil):
    """Recolecta posibles numeros (solo digitos) desde codigo de barras y OCR.
    Devuelve (candidates, debug): candidates es una lista de tuplas
    (digitos, fuente); debug es un dict con detalles crudos para diagnostico
    (que se puede mostrar en pantalla si la deteccion automatica falla)."""
    candidates = []
    debug = {
        "zbar_disponible": ZBAR_AVAILABLE,
        "zbar_lecturas": [],
        "zbar_error": None,
        "ocr_disponible": TESSERACT_AVAILABLE,
        "ocr_lineas": [],
        "ocr_error": None,
    }

    if ZBAR_AVAILABLE:
        try:
            img = np.array(image_pil.convert("L"))
            for result in zbar_decode(img):
                data = result.data.decode("utf-8", errors="ignore").strip()
                debug["zbar_lecturas"].append(data)
                digits = re.sub(r"\D", "", data)
                if len(digits) >= 10:
                    candidates.append((digits, "Codigo de barras/QR"))
        except Exception as e:
            debug["zbar_error"] = f"{type(e).__name__}: {e}"

    if TESSERACT_AVAILABLE:
        processed = _preprocess_for_ocr(image_pil)
        configs = [
            "--psm 6 -c tessedit_char_whitelist=0123456789",
            "--psm 11 -c tessedit_char_whitelist=0123456789",
        ]
        for config in configs:
            try:
                text = pytesseract.image_to_string(processed, config=config)
            except Exception as e:
                debug["ocr_error"] = f"{type(e).__name__}: {e}"
                continue
            for line in text.splitlines():
                line = line.strip()
                if line:
                    debug["ocr_lineas"].append(line)
                digits = re.sub(r"\D", "", line)
                if len(digits) >= 10:
                    candidates.append((digits, "OCR"))

    return candidates, debug


def detect_codes(image_pil):
    """Detecta el ICCID y, si existe, el IMEI/REIF (otro numero largo) en la
    foto. Devuelve (iccid, iccid_metodo, imei, imei_metodo, debug)."""
    image_pil = _normalize_image(image_pil)
    candidates, debug = collect_candidates(image_pil)
    if not candidates:
        return None, "Manual", None, "Manual", debug

    iccid = None
    iccid_metodo = "Manual"
    imei = None
    imei_metodo = "Manual"

    # 1) ICCID "clasico" (empieza en 89)
    for digits, fuente in candidates:
        m = ICCID_REGEX.search(digits)
        if m:
            iccid = m.group(0)
            iccid_metodo = fuente
            break

    # 2) Si no hay ICCID clasico, usar cualquier corrida de 18-20 digitos
    if not iccid:
        for digits, fuente in candidates:
            m = DIGIT_RUN_REGEX.search(digits)
            if m:
                iccid = m.group(0)
                iccid_metodo = fuente
                break

    # 3) IMEI/REIF: otro numero largo (14-20 digitos) distinto del ICCID
    for digits, fuente in candidates:
        if iccid and (digits == iccid or iccid in digits):
            continue
        if len(digits) >= 14:
            imei = digits[:20]
            imei_metodo = fuente
            break

    return iccid, iccid_metodo, imei, imei_metodo, debug


def build_excel_with_photos(df_full):
    """Genera el Excel con ICCID, IMEI, tienda, fecha, metodo y una miniatura
    de la foto de cada chip."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario SIM"

    headers = ["ICCID", "IMEI", "Tienda", "Fecha y Hora de Captura", "Metodo", "Foto"]
    ws.append(headers)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions[get_column_letter(6)].width = 14

    for i, row in enumerate(df_full.itertuples(index=False), start=2):
        ws.cell(row=i, column=1, value=row.ICCID)
        ws.cell(row=i, column=2, value=row.IMEI)
        ws.cell(row=i, column=3, value=row.Tienda)
        ws.cell(row=i, column=4, value=getattr(row, "_3"))
        ws.cell(row=i, column=5, value=row.Metodo)
        ws.row_dimensions[i].height = 70

        foto_path = row.FotoPath
        if foto_path and os.path.exists(foto_path):
            try:
                xl_img = XLImage(foto_path)
                xl_img.width = THUMB_PX
                xl_img.height = THUMB_PX
                ws.add_image(xl_img, f"F{i}")
            except Exception:
                ws.cell(row=i, column=6, value="(no se pudo insertar)")
        else:
            ws.cell(row=i, column=6, value="(sin foto)")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ----------------------------------------------------------------------
# UI
# ----------------------------------------------------------------------
st.set_page_config(page_title="Inventario Chips SIM", page_icon="Chip", layout="centered")

conn = get_connection()

defaults = {
    "detected_code": "",
    "detected_metodo": "Manual",
    "detected_imei": "",
    "detected_imei_metodo": "Manual",
    "captured_image_bytes": None,
    "input_key": 0,
    "camera_key": 0,
    "is_admin": False,
    "tienda_seleccionada": None,
    "last_debug": None,
}
for key, default in defaults.items():
    if key not in st.session_state:
        st.session_state[key] = default

top_col1, top_col2 = st.columns([5, 1])
with top_col1:
    st.title("Inventario de Chips SIM")
with top_col2:
    st.write("")
    with st.popover("Cuenta", use_container_width=True):
        if st.session_state.is_admin:
            st.success("Administrador conectado")
            if st.button("Cerrar sesion", use_container_width=True):
                st.session_state.is_admin = False
                st.rerun()
        else:
            st.caption("Acceso de administrador")
            admin_pwd = st.text_input("Contrasena", type="password", key="admin_pwd_field")
            if st.button("Iniciar sesion", use_container_width=True):
                if not ADMIN_PASSWORD:
                    st.error("Contrasena de administrador no configurada (ver README).")
                elif admin_pwd == ADMIN_PASSWORD:
                    st.session_state.is_admin = True
                    st.rerun()
                else:
                    st.error("Contrasena incorrecta.")

total = conn.execute("SELECT COUNT(*) FROM chips").fetchone()[0]
st.metric("Total de chips registrados", total)

if not ZBAR_AVAILABLE or not TESSERACT_AVAILABLE:
    faltantes = []
    if not ZBAR_AVAILABLE:
        faltantes.append("lectura de codigo de barras/QR (zbar)")
    if not TESSERACT_AVAILABLE:
        faltantes.append("OCR (tesseract)")
    st.warning(
        "No estan disponibles: " + " y ".join(faltantes) +
        ". Puedes seguir usando la app con captura/edicion manual del ICCID. "
        "Revisa el README para instalar las dependencias de sistema."
    )

st.divider()

if not st.session_state.tienda_seleccionada:
    st.subheader("0. Selecciona tu tienda")
    tienda_choice = st.selectbox(
        "Tienda",
        TIENDAS,
        index=None,
        placeholder="Elige una tienda...",
        key="tienda_selectbox",
    )
    if st.button("Confirmar tienda", type="primary", disabled=(tienda_choice is None)):
        st.session_state.tienda_seleccionada = tienda_choice
        st.rerun()
    st.info(
        "Selecciona tu tienda para comenzar a escanear. Esta eleccion queda fija "
        "durante tu sesion (no se puede cambiar sin recargar la pagina)."
    )
else:
    st.caption(f"Tienda seleccionada: **{st.session_state.tienda_seleccionada}**")

    st.subheader("1. Escanear")
    st.caption(
        "En celular: usa la camara trasera, sostenla DE FRENTE al codigo (no en "
        "angulo/inclinada), acercate para que los numeros llenen bien el encuadre, "
        "y procura buena luz sin brillos ni sombras. "
        "La app intenta leer el ICCID y el IMEI/REIF en la misma foto."
    )
    img_file = st.camera_input(
        "Toma una foto del ICCID / codigo de barras del chip",
        key=f"camera_{st.session_state.camera_key}",
    )

    if img_file is not None:
        img_bytes = img_file.getvalue()
        image = Image.open(io.BytesIO(img_bytes))
        with st.spinner("Analizando imagen..."):
            iccid_code, iccid_metodo, imei_code, imei_metodo, debug_info = detect_codes(image)
        if iccid_code:
            st.success(f"ICCID detectado ({iccid_metodo}): {iccid_code}")
        else:
            st.warning("No se detecto el ICCID automaticamente. Ingresalo manualmente abajo.")
        if imei_code:
            st.info(f"IMEI/REIF detectado ({imei_metodo}): {imei_code}")
        st.session_state.detected_code = iccid_code or ""
        st.session_state.detected_metodo = iccid_metodo
        st.session_state.detected_imei = imei_code or ""
        st.session_state.detected_imei_metodo = imei_metodo
        st.session_state.captured_image_bytes = img_bytes
        st.session_state.last_debug = debug_info

    if st.session_state.last_debug:
        d = st.session_state.last_debug
        with st.expander("Ver diagnostico de lectura (util si no detecta nada)"):
            st.write(f"Lectura de codigo de barras disponible: {d['zbar_disponible']}")
            if d["zbar_error"]:
                st.error(f"Error al leer codigo de barras: {d['zbar_error']}")
            st.write(f"Codigos de barras detectados en la foto: {len(d['zbar_lecturas'])}")
            for lectura in d["zbar_lecturas"]:
                st.code(lectura)
            st.divider()
            st.write(f"OCR disponible: {d['ocr_disponible']}")
            if d["ocr_error"]:
                st.error(f"Error en OCR: {d['ocr_error']}")
            st.write(f"Lineas de texto leidas por OCR: {len(d['ocr_lineas'])}")
            for linea in d["ocr_lineas"]:
                st.code(linea)

    iccid_input = st.text_input(
        "ICCID (verifica o corrige antes de guardar)",
        value=st.session_state.detected_code,
        max_chars=22,
        key=f"iccid_input_{st.session_state.input_key}",
    )
    imei_input = st.text_input(
        "IMEI / REIF (opcional, verifica o corrige antes de guardar)",
        value=st.session_state.detected_imei,
        max_chars=22,
        key=f"imei_input_{st.session_state.input_key}",
    )

    col1, col2 = st.columns(2)
    with col1:
        guardar = st.button("Guardar registro", use_container_width=True, type="primary")
    with col2:
        limpiar = st.button("Limpiar", use_container_width=True)

    if guardar:
        clean = re.sub(r"\D", "", iccid_input)
        clean_imei = re.sub(r"\D", "", imei_input) or None
        if not clean:
            st.error("Ingresa un ICCID valido.")
        elif len(clean) < 15:
            st.error("El ICCID parece incompleto (muy corto). Verifica la foto o el texto.")
        elif iccid_exists(conn, clean):
            st.error(f"Este chip ya fue escaneado antes: {clean}")
        else:
            foto_path = None
            if st.session_state.captured_image_bytes:
                foto_path = save_photo(clean, st.session_state.captured_image_bytes)
                if foto_path is None:
                    st.warning("No se pudo guardar la foto, pero el registro si se guardo.")
            insert_chip(
                conn, clean, st.session_state.detected_metodo, foto_path,
                st.session_state.tienda_seleccionada, clean_imei,
            )
            st.success(f"Chip guardado: {clean}")
            st.session_state.detected_code = ""
            st.session_state.detected_imei = ""
            st.session_state.captured_image_bytes = None
            st.session_state.last_debug = None
            st.session_state.input_key += 1
            st.session_state.camera_key += 1
            st.rerun()

    if limpiar:
        st.session_state.detected_code = ""
        st.session_state.detected_imei = ""
        st.session_state.captured_image_bytes = None
        st.session_state.last_debug = None
        st.session_state.input_key += 1
        st.session_state.camera_key += 1
        st.rerun()

st.divider()
st.subheader("2. Chips escaneados")

df_full = get_all(conn)
df_display = df_full.drop(columns=["FotoPath"]).copy()
df_display.insert(
    len(df_display.columns), "Foto",
    df_full["FotoPath"].apply(lambda p: "Si" if p and os.path.exists(p) else "-")
)
st.dataframe(df_display, use_container_width=True, hide_index=True)

st.divider()
st.subheader("3. Exportar")

if not st.session_state.is_admin:
    st.info(
        "La descarga del Excel esta disponible solo para el administrador. "
        "Inicia sesion desde el boton 'Cuenta' arriba a la derecha."
    )
elif len(df_full) > 0:
    with st.spinner("Generando Excel con fotos..."):
        excel_buffer = build_excel_with_photos(df_full)
    st.download_button(
        label="Descargar Excel (incluye fotos)",
        data=excel_buffer,
        file_name=f"inventario_sim_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.caption(
        "Nota: con varios cientos de chips, el Excel puede pesar varios MB y tardar "
        "unos segundos en generarse por las fotos embebidas."
    )
else:
    st.info("Aun no hay chips escaneados.")

with st.expander("Eliminar un registro (correcciones)"):
    if len(df_full) > 0:
        to_delete = st.selectbox("Selecciona el ICCID a eliminar", df_full["ICCID"].tolist())
        if st.button("Eliminar registro seleccionado"):
            delete_chip(conn, to_delete)
            st.rerun()
    else:
        st.caption("No hay registros para eliminar.")

st.caption(
    "Los datos se guardan localmente: la tabla en inventario_sim.db y las fotos en "
    "la carpeta fotos_chips/, en el equipo donde corre la app."
)
